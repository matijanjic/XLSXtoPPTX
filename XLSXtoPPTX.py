
#  A script that takes an existing excel file, reads the required columns that hold the info
# (word, sentence, word audio, sentence audio, sentence picture etc.) and arranges it all into a
# pptx presentation. The use case is pretty specific, my wife needed it for her phd experiment.
# Audio files were generated using another script I made that uses the Google text-to-speech
# (gTTS) library that read the words and sentences from a csv file and exported them as .mp3's

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from collections import defaultdict
from slideshow import *

# returns a dictionary of lists where each letter key has a value that is a list of values from that column
# in the worksheet. There are four kwargs: rowStart, rowEnd, colStart and colEnd which are by default set to 
# "auto", which means the whole table will be considered. Might be problematic when there are gaps in the columns
# or rows, so I'd suggest telling the function explicitly which column and rows are starting and ending ones. 
def getDictFromXlsx(xlsxFile, colList, rowStart = "auto", rowEnd = "auto", colStart = "auto", colEnd = "auto"):

    wb = load_workbook(xlsxFile, data_only=True)
    ws = wb.active

    # if any of the function inputs is requiring the calculation of the start or end points for columns
    # and rows, then execute the code below.
    if rowStart == "auto" or rowEnd == "auto" or colStart == "auto" or colEnd == "auto":
        # checks the whole sheet
        for rows in ws.rows:
            for cell in rows:
                # stops when it finds a non empty cell and asigns that column letter to the firstColumnLetter variable
                if cell.value != None:
                    firstColumnLetter = cell.column_letter
                    firstColumn = cell.column
                    break
        
        # goes through the first column with content and works out first and last row that is not empty
        column = ws[firstColumnLetter]
        cellEmpty = True
        for cell in column:
            if cell.value != None and cell.row == 1:
                firstRow = 1
                cellEmpty = False
            elif cell.value != None and cellEmpty == True:
                firstRow = cell.row
                cellEmpty = False
            elif cell.value == None and cellEmpty == False:
                lastRow = cell.row - 1
                break
        # goes through the first non-empty row and finds the last column not empty
        row = ws[firstRow]
        cellEmpty = True
        for cell in row:
            if cell.column_letter > firstColumnLetter and cell.value == None:
                lastColumn = cell.column - 1
                break
        if rowStart == "auto":
            rowStart = firstRow
        if rowEnd == "auto":
            rowEnd = lastRow
        if colStart == "auto":
            colStart = firstColumn
        if colEnd == "auto":
            colEnd = lastColumn    
        
        

    # goes through the workbook and returns a dictionary with the values and column letters
    wsDict = defaultdict(list)
    for row in range(rowStart, rowEnd + 1):
        for column in range(colStart, colEnd + 1):
            colLetter = get_column_letter(column)
            if colLetter in colList:
                wsDict[colLetter].append(ws.cell(row=row, column=column).value)
    return wsDict


def main():
    # some constants declared here
    xlsxFile = 'excel\spreadsheet_gorilla_learning_pictures_-all.xlsx'
    wordSoundFolder = 'sounds\words\\'
    sentenceSoundFolder = 'sounds\sentences\\'
    pictureFolder = 'images\pictures_learning\\'

    # these could be changed to what ever fits your needs, but are column letters that hold the data for the powerpoint
    wordTextCol = 'F'
    wordSoundCol = 'G'
    sentenceSoundCol = 'I'
    sentencePictureCol = 'J'
    # list of all the columns so they can be searched more easily. If the number of columns were bigger, it would pay off to automate it
    colList = ['F', 'G', 'I', 'J']
    
    # create a new instance of the SlideShow class that takes the width and the height in inches
    # and the and the slide layout (further explained in the python-pptx documentation 
    # https://python-pptx.readthedocs.io/en/latest/user/slides.html)
    # e.g. SlideShow(16, 9, 6) creates a 16x9 presentation file with a blank slide template (layout number 6)
    slideShow = SlideShow(16, 9, 6)

    # use the getDictFromXlsx function that returns a filled out dictionary where the keys are the column letters
    # and the values are lists that contain the data in those columns. 
    xlsxDict = getDictFromXlsx(xlsxFile, colList)
    numberOfRows = len(list(xlsxDict.values())[0])
    print(numberOfRows)
    # -- MAIN SLIDE LAYOUT --#

    # for each row create a following slide layout:
    for i in range(numberOfRows):
        print(i)
        slideShow.addSlide()
        slideShow.addText(80, 'X', 4, 1)
        slideShow.addSlide()
        slideShow.addText(44, xlsxDict[wordTextCol][i], 4, 1)
        slideShow.addSound(wordSoundFolder + xlsxDict[wordSoundCol][i])
        slideShow.addSlide()
        slideShow.addPicture(pictureFolder + xlsxDict[sentencePictureCol][i], 400)
        slideShow.addSound(sentenceSoundFolder + xlsxDict[sentenceSoundCol][i])
    # and save it
    slideShow.save('test.pptx')


if __name__ == '__main__':
    main()
